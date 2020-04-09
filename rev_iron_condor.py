"""
Rev iron condor analyser
Written by: Peter Agalakov
version: v0.42(2020-april-09)

V0.42(2020-april-09)
*Use last price instead of Mid
*Added a filter to output of 150
*Fixed market time

v0.41 (2020-march-31)
*Created individual legs for multiple strategies

v0.4 (2020-march-31)
*extracted all the parameters to be global
*cleaned up and refactored some code to prerp for GUI

v0.3(2020-march-27)
*Added scheduler to take in data ever 30 min
*Added exporting options to excel sheet

v0.2(2020-march-26)
*Added Open Interest and Volume to displayed Dataframe
*Integrated some functionality into functions
*Added for loop so can return multiple expiration dates
*Also return the min Cost from Dataframe and the time for late graph use

v0.1(2020-march-20)
----
Initial development phase
"""

import numpy as np
import pandas as pd
from yahoo_fin.options import *
from yahoo_fin.stock_info import *
from datetime import datetime
import schedule
import time
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# noinspection PyUnresolvedReferences
def get_min_price(p, c, s, legs):
    """4 loops nested in each other, first loop selects the items down the
    dataframe list. The other loops verify that the strategy suggested is
    possible since not all strike prices have the necessary volume. Finally
    if we mange to get to the last leg, everything is appended to a list.
    *** Not really happy with this function but I can't really think right
    now of a way to make it better and since it works fine i will leave it
    as is ***"""
    gen_price_slot = []
    leg1_list = []
    leg2_list = []
    leg3_list = []
    leg4_list = []
    v = []
    volume = 0

    for leg in legs:
        if leg[1] == 'put':
            leg.append(p)
        elif leg[1] == 'call':
            leg.append(c)

    for i in range(len(leg1[2])):
        leg1_strike_price = leg1[2].loc[i, 'Strike']
        leg1_price = leg1[2].loc[i, 'Last Price']
        volume += int(leg1[2].loc[i, 'Volume'])
        for l in range(len(leg2[2])):
            if leg1_strike_price + s[0] == leg2[2].loc[l, 'Strike']:
                leg2_strike_price = leg2[2].loc[l, 'Strike']
                leg2_price = leg2[2].loc[l, 'Last Price']
                volume += int(leg2[2].loc[l, 'Volume'])
                for m in range(len(leg3[2])):
                    if leg1_strike_price + s[0] + s[1] == leg3[2].loc[m, 'Strike']:
                        leg3_strike_price = leg3[2].loc[m, 'Strike']
                        leg3_price = leg3[2].loc[m, 'Last Price']
                        volume += int(leg3[2].loc[m, 'Volume'])
                        for n in range(len(leg4[2])):
                            if leg1_strike_price + s[0] + s[1] + s[2] == \
                                    leg4[2].loc[n, 'Strike']:
                                leg4_strike_price = leg4[2].loc[n, 'Strike']
                                leg4_price = leg4[2].loc[n, 'Last Price']
                                volume = int(leg4[2].loc[n, 'Volume'])
                                value = gen_price_value(leg1_price,
                                                        leg2_price,
                                                        leg3_price,
                                                        leg4_price,
                                                        legs)
                                gen_price_slot.append(value)
                                leg1_list.append(leg1_strike_price)
                                leg2_list.append(leg2_strike_price)
                                leg3_list.append(leg3_strike_price)
                                leg4_list.append(leg4_strike_price)
                                v.append(volume / 4)

    return gen_price_slot, leg1_list, leg2_list, leg3_list, leg4_list, v


def gen_price_value(l1_p, l2_p, l3_p, l4_p, legs):
    value = (legs[0][0] * l1_p) + (legs[1][0] * l2_p) + (legs[2][0] *
                                                           l3_p) \
            + (legs[3][0] * l4_p)
    value = np.around(value, decimals=2)
    return value


def filter_blank(df, col):
    """Filter empty spaces"""
    df_filter = df[col] != '-'
    return df[df_filter]


def filter_volume(df):
    """Filter by volume, default min 50 volume"""
    global volume_filter
    vol_filter = pd.to_numeric(df['Volume']) > volume_filter
    return df[vol_filter]


def get_df(o_df):
    """ Function that filters, converts and rearranges the original
    data frame"""
    o_df = o_df[['Strike', 'Last Price', 'Volume']]
    o_df = filter_blank(o_df, 'Volume')
    o_df = filter_volume(o_df)
    o_df = o_df.reset_index()
    return o_df


def f_df_func(strategy, stock, exp_dates, legs):
    for exp_date in exp_dates:
        sheet_name = str(exp_date[0:4]) + str(exp_date[5:7]) + str(exp_date[
                                                                   8:10])
        # Get put price DataFrame and filter it down to get the necessary
        # columns
        puts_price = get_df(get_puts(stock, exp_date))

        # Get put price DataFrame and filter it down to get the necessary
        # columns
        calls_price = get_df(get_calls(stock, exp_date))

        # Use get_min_price function to obtain a DataFrame with all the price
        # for each leg + Volume + Open interest
        table1, table2, table3, table4, table5, table6 = get_min_price(
            puts_price, calls_price, strategy, legs)

        data = {'Cost': table1, 'Sell_puts': table2, 'Buy_puts': table3,
                'Buy_calls': table4, 'Sell_calls': table5, 'Volume': table6}

        data_frame = pd.DataFrame(data)
        data_frame = filter_volume(data_frame)
        df2 = pd.DataFrame([data_frame['Cost'].max(), datetime.now()])
        df2 = df2.transpose()
        append_df_to_excel('output.xlsx', data_frame,
                           sheet_name=stock + sheet_name)
        append_df_to_excel('output.xlsx', df2,
                           sheet_name=stock + sheet_name + "_Min_cost",
                           index=False, header=None)
        print("Data was added to excel file for:" + stock + " @ " + str(
            datetime.now()))


# Setup-------------------------------------
# Display all rows and columns from DataFrame
pd.set_option('display.max_columns', None)
pd.set_option("max_rows", None)
# Suppresses scientific notation when filling our lists
np.set_printoptions(suppress=True)
now = datetime.now()
day = now.strftime("%A")
s_s_m = (now - now.replace(hour=0, minute=0, second=0,
                           microsecond=0)).total_seconds()
# -----------------------------------------

# Constant parameter to be modified by user
# -----------------------------------------
"""The ticket or the underlying stock"""
stock = 'spy'
"""
The interval desired for each leg of the iron condor in dollars.
ex: 10/20/10 -->> 240/250/270/280
"""
strategy = [5, 10, 5]
variation = [2, 2, 2]
"""Enter the list of expiration dates for contracts """
exp_dates = ['2020/05/15']  # string format 'yyyy/mm/dd'

"""The time interval variable in minutes that the data is collected"""
timer = 15  # in minutes
"""THe days the stock market is open, if trading on only specific days of 
the week this can be modified to ignore other days of the week."""
open_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

volume_filter = 150

"""Buy or sell the current legs (1 = sell, -1 = buy)"""
leg1 = [1, 'put']
leg2 = [-1, 'put']
leg3 = [-1, 'call']
leg4 = [1, 'call']

legs = [leg1, leg2, leg3, leg4]
# -----------------------------------------


schedule.every(timer).minutes.do(f_df_func, strategy, stock, exp_dates,
                                 legs)

while True:
    now = datetime.now()
    day = now.strftime("%A")
    s_s_m = (now - now.replace(hour=0, minute=0, second=0,
                               microsecond=0)).total_seconds()
    # 34200 : 09h30
    # 57600 : 16h00
    if 34200 < s_s_m < 57600 and day in open_days:
        schedule.run_pending()
        time.sleep(1)
    else:
        print("Markets are now closed, the tracker is on stand-by ")
        time.sleep(900)
